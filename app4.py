import streamlit as st
import pandas as pd
from datetime import date, datetime, time, timedelta
import os
from openpyxl import load_workbook

# ============================================
#           PAGE CONFIG (TITLE + ICON)
# ============================================
st.set_page_config(
    page_title="MoodFlow Planner",
    page_icon="✨",
    layout="wide"
)

# Load external CSS for custom UI
with open("theme.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


# ============================================
#               SIDEBAR MENU
# ============================================
st.sidebar.title("✨ MoodFlow Menu")
page = st.sidebar.radio(
    "Go to:",
    ["Dashboard", "Add Tasks", "Daily Plan", "Mood History"]
)

mood_color_map = {
    "Happy": "#FFC107",
    "Sad": "#2196F3",
    "Stressed": "#F44336",
    "Excited": "#FF5722",
    "Neutral": "#9E9E9E",
}


# ============================================
#             GLOBAL STATE INIT
# ============================================
if "tasks" not in st.session_state:
    st.session_state.tasks = []

if "mood" not in st.session_state:
    st.session_state.mood = "Neutral"


# ============================================
#                 HERO HEADER
# ============================================
def hero_section():
    st.markdown("""
    <div class='hero-container'>
        <h1 class='hero-title'>🌟 MoodFlow Planner</h1>
        <p class='hero-subtitle'>Schedule smarter. Flow better. Feel balanced.</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================
#         TIME CALCULATION FOR TASK ORDERING
# ============================================
def compute_schedule_times(tasks, start_time):
    schedule_with_times = []
    current_time = datetime.combine(date.today(), start_time)

    for t in tasks:
        start = current_time
        end = start + timedelta(hours=t["duration_hours"])

        schedule_with_times.append({
            **t,
            "start_time": start.strftime("%H:%M"),
            "end_time": end.strftime("%H:%M")
        })

        current_time = end

    return schedule_with_times


# ============================================
#             DASHBOARD PAGE
# ============================================
if page == "Dashboard":
    hero_section()

    st.subheader("✨ Your daily mood")
    chosen_mood = st.selectbox(
        "How are you feeling today?",
        ["Happy", "Sad", "Stressed", "Excited", "Neutral"],
        index=["Happy", "Sad", "Stressed", "Excited", "Neutral"]
        .index(st.session_state.mood),
    )
    st.session_state.mood = chosen_mood

    st.markdown(
        f"<div class='mood-tag' style='background:{mood_color_map[chosen_mood]}'>"
        f"{chosen_mood}</div>",
        unsafe_allow_html=True
    )

    st.write("Navigate using the sidebar to add tasks or generate your daily plan.")


# ============================================
#             ADD TASK PAGE
# ============================================
elif page == "Add Tasks":
    hero_section()
    st.subheader("➕ Add a Task")

    with st.form("task_form"):
        task_name = st.text_input("Task Name")
        duration = st.number_input(
            "Duration (hours)", min_value=0.1, max_value=10.0, value=1.0
        )
        difficulty = st.selectbox("Difficulty", ["Easy", "Medium", "Hard"])
        submitted = st.form_submit_button("Add Task")

        if submitted:
            if task_name.strip() == "":
                st.error("Task name cannot be empty.")
            else:
                st.session_state.tasks.append({
                    "task_name": task_name,
                    "duration_hours": duration,
                    "difficulty": difficulty
                })
                st.success(f"Task '{task_name}' added!")


    st.subheader("📋 Current Tasks")
    if st.session_state.tasks:
        for t in st.session_state.tasks:
            st.markdown(
                f"""
                <div class="task-card">
                    <h3>{t['task_name']}</h3>
                    <p>⏱ Duration: {t['duration_hours']} hours</p>
                    <p>🔥 Difficulty: {t['difficulty']}</p>
                </div>
                """,
                unsafe_allow_html=True
            )
    else:
        st.info("No tasks added yet.")


# ============================================
#             DAILY PLAN PAGE
# ============================================
elif page == "Daily Plan":
    hero_section()
    st.subheader("🗓 Generate Your Day Plan")

    start_time_input = st.time_input("Start your day at:", value=time(9, 0))

    if st.button("Generate My Plan ✨"):

        if not st.session_state.tasks:
            st.warning("Add at least one task first!")
        else:
            scheduled_tasks = compute_schedule_times(
                st.session_state.tasks, start_time_input
            )

            st.subheader("📝 Your Schedule")
            for i, t in enumerate(scheduled_tasks):
                st.markdown(
                    f"""
                    <div class="timeline-card">
                        <h3>{i+1}. {t['task_name']}</h3>
                        <p>⏱ {t['start_time']} → {t['end_time']}</p>
                        <p>{t['duration_hours']} h | Difficulty: {t['difficulty']}</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

            # SAVE TO EXCEL
            df_rows = [{
                "date": date.today().isoformat(),
                "mood": st.session_state.mood,
                "task_name": t["task_name"],
                "duration_hours": t["duration_hours"],
                "difficulty": t["difficulty"],
                "scheduled_order": i + 1,
                "start_time": t["start_time"],
                "end_time": t["end_time"],
                "motivational_msg": "Keep going, small steps are fine!"
            } for i, t in enumerate(scheduled_tasks)]

            df = pd.DataFrame(df_rows)

            filename = "data/tasks.xlsx"
            os.makedirs("data", exist_ok=True)

            try:
                if not os.path.exists(filename):
                    df.to_excel(filename, index=False, sheet_name="Tasks")
                else:
                    book = load_workbook(filename)
                    with pd.ExcelWriter(filename, engine="openpyxl", mode="a",
                                        if_sheet_exists="overlay") as writer:
                        startrow = book["Tasks"].max_row
                        df.to_excel(writer, index=False, header=False,
                                    sheet_name="Tasks", startrow=startrow)
                st.success("Schedule saved to tasks.xlsx")
            except Exception as e:
                st.error(f"Saving failed: {e}")


# ============================================
#             MOOD HISTORY PAGE
# ============================================
elif page == "Mood History":
    hero_section()
    st.subheader("📊 Mood Tracking")

    filename = "data/tasks.xlsx"
    if os.path.exists(filename):
        try:
            df_history = pd.read_excel(filename)
            mood_counts = df_history["mood"].value_counts()
            st.bar_chart(mood_counts)
        except Exception as e:
            st.warning(f"Could not read history: {e}")
    else:
        st.info("No data yet. Generate a plan first!")
